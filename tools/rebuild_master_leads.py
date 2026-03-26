"""
rebuild_master_leads.py
Rebuilds master_leads.xlsx from three CSV data sources:
  1. Lead Data CSV    – Dubai clinic contacts (Allianzabiz - Lead Data)
  2. Company Details  – Dubai clinic company info with AI/pain-point data
  3. Lead Set1 CSV    – International medspa/dental contacts

Output sheets
  - Leads     : every person-level record, all social fields + "Company Data Available" flag
  - Companies : every company-level record + "Lead Count" (how many leads map to it)
"""

import sys
import io
import re
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# RAW DATA (pasted inline so the script is self-contained)
# ---------------------------------------------------------------------------

LEAD_DATA_CSV = """Company ,Website,Profile,Current Stage,Name,First Name,Last Name,Personal Email,Email,cc,Phone Number,LinkedIn ID,Instagram ID,Facebook,Thread,Notion URL,Notion ID,Instantly ID
Skin111,skin111.com,CEO and Founder,,Dr. Shahram Nabili,Dr. Shahram,Nabili,,skin111clinic@skin111.com,,,https://www.linkedin.com/in/shahram-nabili-b30a2527b/,,,,https://www.notion.so/Dr-Shahram-Nabili-2e8931af353d8179a2c3d6f5ca283c7e,2e8931af353d8179a2c3d6f5ca283c7e,2e8931af-353d-8179-a2c3-d6f5ca283c7e
,,COO,,Supreet Kaur,Supreet,Kaur,joycranberry@gmail.com,supreet@skin111.com,,971 55 548 7793,https://www.linkedin.com/in/supreet-kaur-64298538,,,,https://www.notion.so/Supreet-Kaur-2ef931af353d80c18c62ed91e4f60133,,
,,,,Dr Eman Abdalla,,,,,,,,https://www.instagram.com/dr_eman_abdalla/,,,,
Biolite Clinic Dubai,https://www.biolitedubai.com/,CEO and Founder,,Mona Mirza,Mona,Mirza,,info@biolitedubai.com,,971 50 467 0809,https://www.linkedin.com/in/mona-mirza-439b3112/,https://www.instagram.com/monasyedmirza/?hl=en,,,https://www.notion.so/Mona-Mirza-2e5931af353d815a800bdbd5997b5bf0,,
Eden Aesthics,https://www.edenderma.com/,MD and Owner,,Dr. Farshad Zadeh,Dr. Farshad,Zadeh,,contact@edenderma.com,,,,https://www.instagram.com/dr.farshadzadeh/,https://www.facebook.com/profile.php?id=61565775434275&ref=_xav_ig_profile_page_web#,,https://www.notion.so/Dr-Farshad-Zadeh-2ef931af353d80d781a9c0969eff6b50,2ef931af353d80d781a9c0969eff6b50,
,,Owner,,Christian Forstner,Christian,Forstner,,christian.forstner@edenderma.com,,,https://www.linkedin.com/in/christian-forstner-dubai/,https://www.instagram.com/christian_forstner/,https://www.facebook.com/Christian.Forstner75,,https://www.notion.so/Christian-Forstner-2ef931af353d8021a193e18111cd3d40,2ef931af353d8021a193e18111cd3d40,
Hortman Clinics,https://hortmanclinics.com/,Founder and Owner,,Anastasiia Hortman,Anastasiia,Hortman,,anastasiia@hortmanclinics.com,,,https://www.linkedin.com/in/anastasiia-hortman/,https://www.instagram.com/anastasiia.hortman/?hl=en,,,https://www.notion.so/Anastasiia-Hortman-2ef931af353d8090bfc8c336a2ac46bb,2ef931af353d8090bfc8c336a2ac46bb,
Premium Cosmetic Laser Center,gopremium.ae,MD and Partner,,Hanadi Al-Hariri,Dr. Hanadi,AI Hariri,,hanadi.hariri@gopremium.ae,dr.hanadi@gopremium.ae,971 56 641 8805,https://www.linkedin.com/in/hanadi-al-hariri-14287059/,https://www.instagram.com/hanadi.hariri/,,,,
Medrose Medical Center,https://medrosemedicalcenter.com/,MD and Owner,,Dr.Nadjia Guergour,Dr.Nadjia ,Guergour,,medrosemedical@gmail.com,,,,https://www.instagram.com/dr_nadjiaguergour/,https://www.facebook.com/profile.php?id=100093123412691,,,
Inject Medispa,https://injectmedispa.com/,Medical Director and Founder,,Dr Sarah Coughlan,Dr Sarah ,Coughlan,,drcoughlan@injectmedispa.com,,971 58 585 5520,https://www.linkedin.com/in/dr-sarah-coughlan-a11b002ba/,https://www.instagram.com/drsarahcoughlan,,,,
The Reformery Clinic,https://thereformeryclinic.com/,Managing Director and Co-Founder,,Dr. Pegah Nehzati,Dr. Pegah,Nehzati,,pegah@thereformeryclinic.com,,,https://www.linkedin.com/in/pegah-nehzati-2ba433319/,https://www.instagram.com/dr.pegah.nehzati,,,,
,,Co-Founder,,"Hadi H. Bafi",Hadi,Bafi,,hadi@thereformeryclinic.com,,,https://www.linkedin.com/in/hadi-bafi-7364b121/,https://www.instagram.com/hadi.bafi/,,https://www.threads.com/@hadi.bafi,,"""

COMPANY_DETAILS_CSV = """Clinic / MedSpa Name,Core Services,Positioning / Notes,Website,Why It's Considered Luxury,Known For,Reddit URL (Source),Reddit Context (Buyer Pain Points),Instagram Handle,Engagement,AI Opportunity,Why (AI Fit Logic),Personalized First-Line Opener,Retainer Potential,Why This Clinic Can Pay This,No. of leads Data we have?
EDEN Aesthetics Clinic,"Botox, fillers, lasers",Premium aesthetic clinic,https://edenaestheticsclinic.com,"Award-winning, high-end interiors, concierge feel & personalized care","Botox, fillers, longevity, laser & wellness services",https://www.reddit.com/r/Dubai/comments/medspa_recommendations,"Users ask about trust, pricing transparency, post-treatment follow-ups",@edenaestheticsclinic,Medium,AI Lead Follow-up + AI Consultation,High intent leads need instant responses + reassurance before booking,"I came across EDEN Aesthetics and loved how premium and thoughtfully curated the entire patient experience feels.",High ($2.5k-4k/mo),Luxury positioning + high-intent clientele,
Dr. Safa Aesthetic Clinic,"Botox, fillers",Doctor-led boutique,https://drsafa.com,Doctor-led brand with strong credibility,Botox & fillers by certified doctor,https://www.reddit.com/r/Dubai/comments/botox_doctor_dubai,Pain around choosing the right doctor & safety concerns,@drsafaestheticclinic,Medium-High,AI Pre-Consultation Assistant,Filters serious patients & answers safety FAQs before doctor time,"What stood out to me about Dr. Safa's clinic is the strong doctor-led trust you've built in a crowded aesthetics market.",High ($2k-3.5k/mo),Doctor authority + premium consult value,
Biolite Clinic Dubai,"Injectables, wellness",Luxury medspa,https://www.biolitedubai.com,"Prestigious, holistic aesthetic + wellness approach near Burj Al Arab","Cosmetic injectables + holistic medical aesthetic services",https://www.reddit.com/r/Dubai/comments/luxury_medspa_dubai,"High-income users asking 'best luxury clinic regardless of cost'",@bioliteclinicdubai,High,AI Concierge + VIP Scheduling,Matches luxury expectations + white-glove experience 24/7,"Biolite clearly caters to a clientele that values luxury and discretion over price.",Very High ($4k-6k/mo),Celebrity / ultra-luxury clientele,
Derma One Aesthetic Center,"Fillers, skin treatments",Dermatology + aesthetics,https://dermaone.ae,Medical dermatology combined with aesthetics,Skin treatments & clinical dermatology,https://www.reddit.com/r/Dubai/comments/skin_clinic_dubai,Concerns about skin conditions + long response times,@dermaoneclinic,Medium,AI Triage + Appointment Routing,"Reduces front-desk overload, routes cases correctly","Derma One's blend of medical dermatology with aesthetics really sets it apart from typical medspas.",Medium-High ($2k-3k/mo),Clinical depth + steady patient flow,
Glamorous Aesthetic Clinic,"Botox, fillers",Influencer-friendly brand,https://glamorousdubai.ae,Influencer-heavy branding & social proof,Cosmetic injectables,https://www.reddit.com/r/Dubai/comments/influencer_clinic_reviews,Users skeptical about Instagram vs real results,@glamorousclinic,High,AI Review + Trust Builder Bot,"Handles objections, showcases verified results automatically","Glamorous Aesthetic Clinic has built impressive visibility, especially through social proof and influencer traction.",Medium-High ($2k-3k/mo),Strong inbound + marketing spend,
King's College Hospital Aesthetics,"Botox, fillers",Hospital-backed,https://kingscollegehospitaldubai.com,British medical credibility,Med-backed aesthetics,,Trust & safety,@kingscollegehospital,Medium,AI Medical FAQ Bot,Handles complex medical questions,"King's College Hospital brings a rare level of medical credibility into aesthetics.",Very High ($5k-7k/mo),Institutional budgets + compliance,
Premium Cosmetic Laser Center,"Botox, lasers",High-volume clinic,https://premiumcosmeticlaser.com,Focused laser treatments with highly trained specialists,Laser hair removal + gentle laser skincare,https://www.reddit.com/r/Dubai/comments/laser_treatment_recs,"Pain points about pain, pricing and session effectiveness",@premiumcosmeticlaser (assumed),Low-Medium,AI Appointment + Follow-Up System,High volume laser leads – automation reduces no-shows & questions,,,
LA Aesthetic Medical Center,"Injectables, contouring",International branding,https://laaestheticcenter.com,Broad non-surgical beauty + rejuvenation services,"Injectables, facials, contouring",https://www.reddit.com/r/Dubai/comments/aesthetics_clinic_recs,Users uncertain about procedure value & package pricing,@laaestheticcenter,Medium,AI Package Comparison Bot,Helps prospects compare treatments and choose based on budget/goal,"LA Aesthetic Center offers a surprisingly wide range of non-surgical options under one roof.",Medium ($1.5k-2.5k/mo),"Broad services, mid-luxury",
Milena Aesthetic Clinic,"Botox, fillers",Boutique luxury,https://milenaclinic.ae,Strong aesthetic presence,Face & body,https://www.reddit.com/r/dubai/comments/1lh29kt/best_aesthetic_clinics_in_dubai/,Recommended for fillers & beauty work,@milenaclinic,High,Excellent,"High inbound DMs, luxury clientele, booking-heavy","Milena Clinic's aesthetic presence is strong and consistent across treatments.",Medium ($1.5k-2.5k/mo),Stable demand,
Kaya Skin Clinic Dubai,"Botox, fillers",National chain,https://www.kaya.in,"Established premium chain, consistent service",Laser & dermatology,https://www.reddit.com/r/Dubai/comments/laser_hair_removal_dubai,"Pain around pricing, packages, and follow-ups",@kayaskinclinic,High,AI CRM + Missed Lead Recovery,High volume – many missed leads = revenue leak,"Kaya's scale and brand recognition in Dubai is impressive – very few clinics operate at that volume.",Very High ($4k-6k/mo),Chain + massive lead volume,
Novomed Aesthetic Clinic,Large medical group,,,Hospital-grade infrastructure with luxury feel,Multi-specialty care,https://www.reddit.com/r/Dubai/comments/multispecialty_clinic_dubai,Repeatedly recommended for Botox & fillers,@novomed,Medium-High,AI Patient Navigation System,Simplifies service selection & reduces admin friction,"Novomed's multi-specialty setup makes it one of the most comprehensive healthcare brands in the region.",High ($3k-5k/mo),Hospital-grade infra + cross-sell,
Molodost Clinic,,Boutique aesthetics,https://molodostclinic.com,,,https://www.reddit.com/r/dubai/comments/1lh29kt/best_aesthetic_clinics_in_dubai/,Mentioned as a good aesthetic option,@molodostclinic,Medium,Strong,"Premium positioning, manual follow-ups likely",,,,
Skin Laundry Dubai,,Medical aesthetics / laser,https://www.skinlaundry.com,,,https://www.reddit.com/r/dubai/comments/1lh29kt/best_aesthetic_clinics_in_dubai/,Praised for tear-trough & precision work,@skinlaundry,High,Excellent,"Global brand, massive inbound, automation-ready",,,,
Seline Clinic,,Doctor-led aesthetics,https://selineclinic.com,,,https://www.reddit.com/r/dubai/comments/1lh29kt/best_aesthetic_clinics_in_dubai/,Dr Leonard Josipovic recommended,@selineclinic,Medium,Strong,"Doctor-led, consultation-heavy workflows",,,,
Avoria Aesthetic Clinic,,Aesthetic clinic,https://avoriaclinic.com,,,https://www.reddit.com/r/dubai/comments/1lh29kt/best_aesthetic_clinics_in_dubai/,Natural-looking filler results,@avoriaclinic,Medium,Strong,Natural-results positioning – high consult volume,,,,
SKIN111 Clinic,"Botox, fillers","Luxury chain, High-end injectables",https://skin111.com,"Multi-location luxury aesthetic hubs with valet parking & boutique vibe","Botox, fillers, high-end dermatology treatments",https://www.reddit.com/r/dubai/comments/1lh29kt/best_aesthetic_clinics_in_dubai/,DHA-certified doctors & safety,@skin111clinic,High,Excellent,"Multi-branch, high DM + WhatsApp load",,,,
Lucia Clinic,,Dermatology & aesthetics,https://lucia-clinic.com,,,https://www.reddit.com/r/dubai/comments/109ue8p/reputable_aesthetics_clinics_in_dubai/,Flagged as reputable & trustworthy,@luciaclinic,Medium,Strong,"Trusted doctors, patient education workflows",,,,
Al Biraa Clinic,,Medical aesthetics,https://albiraaclinic.com,,,https://www.reddit.com/r/dubai/comments/109ue8p/reputable_aesthetics_clinics_in_dubai/,Mentioned for authentic products,@albiraaclinic,Low-Medium,Moderate,"Less marketing polish, ops-heavy",,,,
Mediclinic Enhance,,Hospital-backed aesthetics,https://www.mediclinic.ae/en/corporate/enhance.html,,,https://www.reddit.com/r/dubai/comments/109ue8p/reputable_aesthetics_clinics_in_dubai/,High-quality care at Dubai Mall,@mediclinicme,Medium,Strong,Corporate + appointment complexity,,,,
Aesthetics International,,Doctor-led clinic,https://aestheticsinternational.com,,,https://www.reddit.com/r/dubai/comments/109ue8p/reputable_aesthetics_clinics_in_dubai/,Strong injectable results,@aestheticsinternational,Medium,Strong,"Surgeon-led, high-value consults",,,,
Obagi Clinic Dubai Mall,,Luxury dermatology,https://obagimedical.com,,,https://www.reddit.com/r/dubai/comments/tente1/botox_recommendation/,Positive Botox experiences,@obagimedical,Medium,Strong,"Luxury dermatology, repeat follow-ups",,,,
Proderma Clinic,,Dermatology & injectables,https://proderma.ae,,,https://www.reddit.com/r/dubai/comments/tente1/botox_recommendation/,Mentioned in Botox threads,@prodermaclinic,Low-Medium,Moderate,"Traditional ops, needs education automation",,,,
YugenCare,,Boutique medical clinic,https://yugencare.com,,,https://www.reddit.com/r/dubai/comments/109ue8p/reputable_aesthetics_clinics_in_dubai/,Named as trustworthy option,@yugencare,Low,Moderate,"Boutique clinic, WhatsApp-heavy",,,,
C37 Medical Hub,,Aesthetic medical hub,https://c37.ae,,,https://www.reddit.com/r/UAE/comments/1lh2b28/best_aesthetic_clinics_in_uae/,Praised for natural results,@c37medicalhub,Medium,Strong,Multi-doctor coordination pain,,,,
Silkor Laser & Aesthetic,,,,"Big brand, premium but high-volume",Laser hair removal,https://www.reddit.com/r/Dubai/comments/silkor_reviews,Complaints about wait times & callbacks,@silkordubai,High,AI Appointment + Reminder System,Automation needed to handle scale efficiently,"Silkor's ability to operate at such high volume while maintaining a premium image is rare.",Very High ($4k-6k/mo),Scale + operational pain = urgency,
Lucia Clinic Dubai,,,,"Ultra-luxury, discreet, elite clientele",Cosmetic surgery & aesthetics,https://www.reddit.com/r/Dubai/comments/elite_aesthetic_clinic,"Users want privacy, exclusivity, fast response",@luciaclinicdubai,Medium,AI Private Concierge,AI fits luxury privacy-first communication,"Lucia Clinic feels intentionally discreet and elite – clearly built for a very specific audience.",Very High ($4k-6k/mo),Ultra-high-net-worth clients,
Aesthetica Clinic Dubai,,,,"Boutique, personalized treatments",Injectables & skincare,https://www.reddit.com/r/Dubai/comments/boutique_medspa,Patients seek custom care & detailed explanations,@aestheticaclinic,Medium,AI Education + Consultation Bot,Educates before consult – higher conversion,"I noticed how personalized and boutique Aesthetica Clinic's approach feels compared to larger chains.",Medium ($1.5k-2.5k/mo),"Smaller size, high care standards",
Hortman Clinics,,,,"Personalized patient care with European aesthetic standards",Injectables & advanced aesthetics,,Seeking trusted results with personalized care,@hortmanclinics,Medium,AI Consultation Assistant,High-quality personalized approach maps well to AI pre-consultation screening,,,
Eden Aesthetics Clinic,,,,"Award-winning, high-end interiors, concierge feel","Botox, fillers, laser",,Trust and post-treatment follow-up,@edenaesthetics,Medium,AI Follow-up + Rebooking,Concierge feel can be extended with AI follow-up,,,
Inject Medispa,,,,"Doctor-led medical aesthetic clinic with surgical precision","Injectables, facials",,Safety and doctor credentials,@injectmedispa,Medium,AI Pre-Screening + FAQ,Doctor-led trust building automated pre-consultation,,,
The Reformery Clinic,,,,"Holistic wellness + aesthetics integration","Injectables, wellness",,Holistic results and wellness integration,@thereformeryclinic,Medium,AI Wellness Journey Mapping,Map client wellness goals to treatments using AI,,,
Medrose Medical Center,,,,Medical-grade aesthetics with clinical approach,"Injectables, skin treatments",,Clinical credibility and treatment outcomes,@medrosemedicalcenter,Low-Medium,AI Patient Education Bot,Clinical setting benefits from patient education automation,,,"""

# LEAD_SET1 is now read from disk (data/lead_set1.csv) — see main()
_LEAD_SET1_UNUSED = """First Name,Last Name,Email,Email Verification Status,Job Title,LinkedIn,Twitter,Facebook,Website,City,State,Country,Company,Company Domain,Company Website,Company Industry
David,Roze,david@drroze.com,valid,Founder and CEO,https://www.linkedin.com/in/david-roze-biological-dentist,,https://facebook.com/david.roze.90,https://www.drroze.com,Dubai,,United Arab Emirates,ROZE Bio Health,drroze.com,https://www.drroze.com,Hospitals & Healthcare
Jennifer,Shulstad,jennifer@satinmedspa.com,valid,Founder and CEO,https://www.linkedin.com/in/jennifershulstad,,https://facebook.com/jennifershulstad,http://www.satinmedspa.com,Charlotte,NC,United States,Satin Med Spa,satinmedspa.com,http://www.satinmedspa.com,Medical Practice
Joe,Stanoszek,joe@viomedspa.com,valid,Founder,https://www.linkedin.com/in/joe-stanoszek-7b069051,,https://www.facebook.com/joe.stanoszek,http://www.viomedspa.com,Strongsville,OH,United States,Vio Med Spa,viomedspa.com,http://www.viomedspa.com,Medical Practice
Nuwan,Dayawansa,nuwan@upkeepmedspa.com,valid,Co-Founder,https://www.linkedin.com/in/nuwan-dayawansa-379a18105,,https://www.facebook.com/nuwan.foley,http://upkeepmedspa.com,Dallas,TX,United States,UPKEEP MED SPA,upkeepmedspa.com,http://upkeepmedspa.com,Medical Practice
Eric,Dore,eric@couturemedspa.com,valid,Founder,https://www.linkedin.com/in/ericdore,,,http://couturemedspa.com,Winter Park,FL,United States,Couture Med Spa,couturemedspa.com,http://couturemedspa.com,Medical Practice
Mary,Pardee,drmary@modrnmed.com,valid,Founder,https://www.linkedin.com/in/mary-pardee,,,https://modrnmed.com,Marina del Rey,CA,United States,Modrn Med,modrnmed.com,https://modrnmed.com,Medical Practice
Carnesha,Humphries,carnesha.humphries@viomedspa.com,valid,Franchise Owner,https://www.linkedin.com/in/carnesha-humphries-b32154251,,,http://www.viomedspa.com,Houston,TX,United States,VIO Med Spa,viomedspa.com,http://www.viomedspa.com,Medical Practice
Charmy,Patel,charmy.patel@viomedspa.com,valid,Owner,https://www.linkedin.com/in/charmy-patel-180970300,,,http://www.viomedspa.com,Hendersonville,TN,United States,VIO Med Spa,viomedspa.com,http://www.viomedspa.com,Medical Practice
Laurence,Humphries,laurence.humphries@viomedspa.com,valid,Franchise Owner,https://www.linkedin.com/in/laurence-humphries-8872842,https://twitter.com/lah7770,,http://www.viomedspa.com,Houston,TX,United States,VIO Med Spa,viomedspa.com,http://www.viomedspa.com,Medical Practice
Tracy,Olson,tracy@youthtopiamedspa.com,valid,Owner,https://www.linkedin.com/in/tracywhiteolson,,,http://youthtopiamedspa.com/,Atlanta,GA,United States,Youthtopia Med Spa,youthtopiamedspa.com,http://youthtopiamedspa.com/,Medical Practice
Joanna,Boyer,joanna@fundamentalsmedspa.com,valid,Owner,https://www.linkedin.com/in/joanna-boyer-07793232,,https://facebook.com/joanna.holzerboyer,http://www.fundamentalsmedspa.com,Indianapolis,IN,United States,Fundamentals Med Spa,fundamentalsmedspa.com,http://www.fundamentalsmedspa.com,Medical Practice
Steven,Broughton,drb@lightdentalstudios.com,valid,Founder and CEO,https://www.linkedin.com/in/steven-broughton-1373a517,,,https://linktr.ee/lightdentalstudios,,,United States,Light Dental Studios,lightdentalstudios.com,https://linktr.ee/lightdentalstudios,Hospitals & Healthcare
Anthony,Rosich,anthony.rosich@secretmedspa.com,valid,Chief Financial Officer,https://www.linkedin.com/in/anthony-rosich-613a079,,,http://www.secretmedspa.com,Dallas,TX,United States,It's A Secret Med Spa,secretmedspa.com,http://www.secretmedspa.com,Hair Salons and Cosmetology
Sonia,Cotto,sonia@ketamineclinicsouthflorida.com,valid,"Chief Executive Officer, Founder, Provider",https://www.linkedin.com/in/soniacottocrna,https://twitter.com/i/user/243888128,https://www.facebook.com/vbcrew82,,Fort Lauderdale,FL,United States,Ketamine & Wellness Clinic of South Florida,ketamineclinicsouthflorida.com,https://www.ketamineclinicsouthflorida.com,Mental Health Care
Ryan,Rose,ryan.rose@viomedspa.com,valid,Chief Executive Officer,https://www.linkedin.com/in/ryanrose2,https://www.twitter.com/greenrose77,https://www.facebook.com/ryan.rose.927,,Cleveland,OH,United States,VIO Med Spa,viomedspa.com,http://www.viomedspa.com,Medical Practice
Veronica,Tamaru,veronica@beautyloungemedspa.com,valid,Co-Founder and Chief Operating Officer,https://www.linkedin.com/in/veronica-tamaru-438734212,,,,Los Angeles,CA,United States,Beauty Lounge Med Spa,beautyloungemedspa.com,http://www.beautyloungemedspa.com,Health Wellness & Fitness
Stephanie,Joyce,stephanie@attunemedspa.com,valid,Founder and Chief Executive Officer,https://www.linkedin.com/in/myriamstephaniejoyce,,,,New York,NY,United States,Attune Med Spa,attunemedspa.com,https://www.attunemedspa.com,
Alexis,Renda,alexis@alexislauren.com,valid,Chief Executive Officer,https://www.linkedin.com/in/alexis-renda-b6687623b,,,,Miami,FL,United States,ALEXIS LAUREN,alexislauren.com,https://alexislauren.com,
Mark,Greenspan,mark@beautyfixmedspa.com,valid,Founder and Co-CEO,https://www.linkedin.com/in/mark-greenspan-a254264,,https://www.facebook.com/mark.greenspan.73,https://www.beautyfixmedspa.com/,New York,NY,United States,BeautyFix MedSpa,beautyfixmedspa.com,https://www.beautyfixmedspa.com/,Hair Salons and Cosmetology
Marisa,Martino,marisa@skinneymedspa.com,valid,"Founder, Laser Skin Specialist",https://www.linkedin.com/in/marisa-martino-95226260,,https://www.facebook.com/marisa.martino.9,https://www.skinneymedspa.com/,New York,NY,United States,SKINNEY Medspa,skinneymedspa.com,https://www.skinneymedspa.com/,Hospitals & Healthcare
Jason,Feldman,jason.feldman@idealimage.com,valid,Chief Executive Officer,https://www.linkedin.com/in/jasonlfeldman,,,http://www.idealimage.com,Miami,FL,United States,Ideal Image,idealimage.com,http://www.idealimage.com,Hospitals & Healthcare
Carrie,Hershman,carrie.hershman@ovme.com,valid,Chief Executive Officer,https://www.linkedin.com/in/carrie-hershman-3a9a639a,,,http://www.ovme.com,Charlotte,NC,United States,OVME Aesthetics,ovme.com,http://www.ovme.com,Cosmetics & Personal Care Products
Eric,Schweiger,eschweiger@schweigerderm.com,valid,C.E.O. and Founder,https://www.linkedin.com/in/eric-schweiger-19046714,https://twitter.com/drschweiger,https://facebook.com/eric.schweiger,,New York,NY,United States,Schweiger Dermatology Group,schweigerderm.com,http://www.schweigerderm.com,Hospitals & Healthcare
Clint,Weiler,clint.weiler@milanlaser.com,valid,Chief Executive Officer (CEO),https://www.linkedin.com/in/clintweiler,,https://facebook.com/clint.weiler,http://www.milanlaser.com,Los Angeles,CA,United States,Milan Laser Hair Removal,milanlaser.com,http://www.milanlaser.com,Health Wellness & Fitness
Pat,Phelan,pat@sisuclinic.com,valid,Co-Founder and CEO,https://www.linkedin.com/in/patphelan,https://twitter.com/patphelan,https://facebook.com/phelan1,,Cork,,Ireland,Sisu Clinic,sisuclinic.com,http://www.sisuclinic.com,Hospitals & Healthcare
Karen,Castelletti,karen@everbody.com,valid,"Co-Founder, Chief Technology Officer",https://www.linkedin.com/in/acwaaaon0k4bjf2wkofpusxjznzhooacozk_8x8,,https://facebook.com/karen.castelletti,,Brooklyn,NY,United States,Ever/Body,everbody.com,http://www.everbody.com,Hospitals & Healthcare
Nicci,Levy,nicci@alchemy43.com,valid,Founder and CEO,https://www.linkedin.com/in/nicci-levy-8881056,https://twitter.com/Nicciknows,https://facebook.com/nicci.levy.3,,Beverly Hills,CA,United States,Alchemy 43,alchemy43.com,http://www.alchemy43.com,Consumer Services - General
Gabrielle,Garritano,gabby@jectnyc.com,valid,Founder and CEO,https://www.linkedin.com/in/gabbygarritano,,,,New York,NY,United States,JECT,jectnyc.com,http://jectnyc.com,Health Wellness & Fitness
Mari,Santos,mari@marisclinic.com,valid,Founder and CEO,https://www.linkedin.com/in/mari-santos-a604b51b7,,,,Dubai,,United Arab Emirates,Maris Aesthetic Clinic,marisclinic.com,https://www.marisclinic.com,
Zieda,Sharipova,zieda@ziedasclinic.com,valid,Founder,https://www.linkedin.com/in/zieda-sharipova-aa955a194,,,,,,United Arab Emirates,Zieda Aesthetic Clinic,ziedasclinic.com,http://ziedasclinic.com/,
Dimple,Patel,dimple@glowdubai.ae,valid,Company Owner,https://www.linkedin.com/in/dimple-patel-uae,,,,Dubai,,United Arab Emirates,"Glow Aesthetic, Medical & Dermatology Clinic",glowdubai.ae,https://glowdubai.ae/,
Sanjay,Parashar,sanjay@cocoona.ae,valid,Managing Director,https://www.linkedin.com/in/sanjay-parashar-35216a2,https://twitter.com/i/user/54488091,https://www.facebook.com/drsanjayparashar,,Dubai,,United Arab Emirates,MedArt Clinic,,,
Dean,Vistnes,dvistnes@skinspirit.com,valid,Medical Director and Co-Founder,https://www.linkedin.com/in/dean-vistnes-30a62621,,https://www.facebook.com/dean.vistnes,http://www.skinspirit.com,Palo Alto,CA,United States,SkinSpirit Skincare Clinic & Spa,skinspirit.com,http://www.skinspirit.com,Cosmetics & Personal Care Products
Steven,Welch,steve@restore.com,valid,Co-Founder and Board Member,https://www.linkedin.com/in/livefreeordie,https://twitter.com/welchsteven,https://facebook.com/steven.welch,,Austin,TX,United States,Restore Hyper Wellness,restore.com,https://restore.com/,Health Wellness & Fitness
Amy,Neary,aneary@primeivhydration.com,valid,CEO,https://www.linkedin.com/in/amy-neary-05545ba,,https://facebook.com/amy.b.mullins.16,http://primeivhydration.com,Colorado Springs,CO,United States,Prime IV Hydration & Wellness,primeivhydration.com,http://primeivhydration.com,Fitness & Dance Facilities
Eric,Casaburi,eric@serotonincenters.com,valid,Chief Executive Officer,https://www.linkedin.com/in/ericcasaburi,https://twitter.com/ecretro,https://facebook.com/eric.casaburi,,Orlando,FL,United States,SEROTONIN Anti-Aging Centers,serotonincenters.com,http://www.serotonincenters.com,,
Esther,Fieldgrass,esther@efmedispa.com,valid,Founder,https://www.linkedin.com/in/esther-fieldgrass-1341b585,,,http://efmedispa.com/,,,United Kingdom,EF MEDISPA,efmedispa.com,http://efmedispa.com/,Cosmetics & Personal Care Products
Sam,Cinkir,sam@estemedicalgroup.uk,valid,Founder and CEO,https://www.linkedin.com/in/sam-cinkir,,https://facebook.com/huseyin.cinkir.1,https://www.estemedicalgroup.uk/,,,United Kingdom,Este Medical Group UK,estemedicalgroup.uk,https://www.estemedicalgroup.uk/,,
Mona,Mirza,mona.mirza@biolitedubai.com,risky,CEO,https://www.linkedin.com/in/mona-mirza-439b3112,,,,Dubai,,United Arab Emirates,Biolite Aesthetic Clinic,biolitedubai.com,http://www.biolitedubai.com,Leisure & Hospitality - General"""

# ---------------------------------------------------------------------------
# PARSE HELPERS
# ---------------------------------------------------------------------------

def parse_csv_text(text):
    """Simple CSV parser that handles quoted fields."""
    import csv
    reader = csv.DictReader(io.StringIO(text.strip()))
    return list(reader)

# ---------------------------------------------------------------------------
# NORMALISE COMPANY NAME for matching
# ---------------------------------------------------------------------------

def norm(name):
    if not name:
        return ""
    n = name.strip().lower()
    # Remove common suffixes / punctuation
    n = re.sub(r'[^a-z0-9 ]', '', n)
    n = re.sub(r'\b(clinic|clinics|medspa|med spa|medical center|aesthetics|aesthetic|dubai|uae|ltd|llc|inc)\b', '', n)
    n = re.sub(r'\s+', ' ', n).strip()
    return n

# ---------------------------------------------------------------------------
# STYLES
# ---------------------------------------------------------------------------

HEADER_FILL   = PatternFill("solid", fgColor="1F4E79")
SUB_FILL      = PatternFill("solid", fgColor="2E75B6")
ALT_FILL      = PatternFill("solid", fgColor="D6E4F0")
WHITE_FILL    = PatternFill("solid", fgColor="FFFFFF")
YES_FILL      = PatternFill("solid", fgColor="C6EFCE")
NO_FILL       = PatternFill("solid", fgColor="FFE0E0")
WARN_FILL     = PatternFill("solid", fgColor="FFEB9C")
HEADER_FONT   = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
BODY_FONT     = Font(name="Calibri", size=10)
BOLD_FONT     = Font(name="Calibri", bold=True, size=10)
CENTER        = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT          = Alignment(horizontal="left",   vertical="center", wrap_text=True)
thin          = Side(style="thin", color="BFBFBF")
BORDER        = Border(left=thin, right=thin, top=thin, bottom=thin)

def style_header(cell, sub=False):
    cell.fill   = SUB_FILL if sub else HEADER_FILL
    cell.font   = HEADER_FONT
    cell.alignment = CENTER
    cell.border = BORDER

def style_body(cell, alt=False, fill_override=None):
    cell.fill   = fill_override or (ALT_FILL if alt else WHITE_FILL)
    cell.font   = BODY_FONT
    cell.alignment = LEFT
    cell.border = BORDER

def write_row(ws, row_idx, values, alt=False, fills=None):
    for col_idx, val in enumerate(values, 1):
        c = ws.cell(row=row_idx, column=col_idx, value=val)
        fo = fills.get(col_idx) if fills else None
        style_body(c, alt=alt, fill_override=fo)

# ---------------------------------------------------------------------------
# BUILD LEADS
# ---------------------------------------------------------------------------

def build_leads(lead_data_rows, lead_set1_rows):
    """Returns list of dicts, each representing one lead."""
    leads = []

    # --- Source 1: Lead Data (Dubai clinics) ---
    current_company = ""
    current_website = ""
    for r in lead_data_rows:
        company  = (r.get("Company ") or r.get("Company") or "").strip()
        website  = (r.get("Website") or "").strip()
        if company:
            current_company = company
            current_website = website
        elif not company and current_company:
            company = current_company
            website = current_website

        name       = (r.get("Name") or "").strip()
        first      = (r.get("First Name") or "").strip()
        last       = (r.get("Last Name") or "").strip()
        if not first and not last and name:
            parts = name.split()
            first = parts[0] if parts else name
            last  = " ".join(parts[1:]) if len(parts) > 1 else ""

        leads.append({
            "Source":        "Lead Data",
            "First Name":    first,
            "Last Name":     last,
            "Full Name":     name or f"{first} {last}".strip(),
            "Job Title":     r.get("Profile", ""),
            "Company":       company,
            "Email":         (r.get("Email") or "").strip(),
            "Personal Email":(r.get("Personal Email") or "").strip(),
            "CC Email":      (r.get("cc") or "").strip(),
            "Email Status":  "",
            "Phone":         (r.get("Phone Number") or "").strip(),
            "LinkedIn":      (r.get("LinkedIn ID") or "").strip(),
            "Instagram":     (r.get("Instagram ID") or "").strip(),
            "Facebook":      (r.get("Facebook") or "").strip(),
            "Twitter":       "",
            "Website":       website,
            "City":          "",
            "State":         "",
            "Country":       "United Arab Emirates",
            "Company Domain":"",
            "Company Website": website,
            "Industry":      "",
            "Notion URL":    (r.get("Notion URL") or "").strip(),
            "Instantly ID":  (r.get("Instantly ID") or "").strip(),
            "Current Stage": (r.get("Current Stage") or "").strip(),
        })

    # --- Source 2: Lead Set1 (International) ---
    for r in lead_set1_rows:
        leads.append({
            "Source":        "Lead Set 1",
            "First Name":    r.get("First Name", ""),
            "Last Name":     r.get("Last Name", ""),
            "Full Name":     f"{r.get('First Name','')} {r.get('Last Name','')}".strip(),
            "Job Title":     r.get("Job Title", ""),
            "Company":       r.get("Company", ""),
            "Email":         r.get("Email", ""),
            "Personal Email":"",
            "CC Email":      "",
            "Email Status":  r.get("Email Verification Status", ""),
            "Phone":         "",
            "LinkedIn":      r.get("LinkedIn", ""),
            "Instagram":     "",
            "Facebook":      r.get("Facebook", ""),
            "Twitter":       r.get("Twitter", ""),
            "Website":       r.get("Website", ""),
            "City":          r.get("City", ""),
            "State":         r.get("State", ""),
            "Country":       r.get("Country", ""),
            "Company Domain":r.get("Company Domain", ""),
            "Company Website":r.get("Company Website", ""),
            "Industry":      r.get("Company Industry", ""),
            "Notion URL":    "",
            "Instantly ID":  "",
            "Current Stage": "",
        })

    return leads

# ---------------------------------------------------------------------------
# BUILD COMPANIES
# ---------------------------------------------------------------------------

def build_companies(company_rows):
    companies = []
    seen = set()
    for r in company_rows:
        name = (r.get("Clinic / MedSpa Name") or "").strip()
        if not name or norm(name) in seen:
            continue
        seen.add(norm(name))
        companies.append({
            "Company Name":        name,
            "Website":             r.get("Website", ""),
            "Core Services":       r.get("Core Services", ""),
            "Positioning":         r.get("Positioning / Notes", ""),
            "Why Luxury":          r.get("Why It's Considered Luxury", ""),
            "Known For":           r.get("Known For", ""),
            "Instagram Handle":    r.get("Instagram Handle", ""),
            "Engagement Level":    r.get("Engagement", ""),
            "Reddit URL":          r.get("Reddit URL (Source)", ""),
            "Buyer Pain Points":   r.get("Reddit Context (Buyer Pain Points)", ""),
            "AI Opportunity":      r.get("AI Opportunity", ""),
            "AI Fit Logic":        r.get("Why (AI Fit Logic)", ""),
            "Personalized Opener": r.get("Personalized First-Line Opener", ""),
            "Retainer Potential":  r.get("Retainer Potential", ""),
            "Why They Can Pay":    r.get("Why This Clinic Can Pay This", ""),
        })
    return companies

# ---------------------------------------------------------------------------
# CROSS-REFERENCE
# ---------------------------------------------------------------------------

def cross_reference(leads, companies):
    """Add 'Company Data Available' to leads and 'Lead Count' to companies."""
    comp_norms = {norm(c["Company Name"]): c["Company Name"] for c in companies}

    for lead in leads:
        cn = norm(lead["Company"])
        matched = cn in comp_norms
        lead["Company Data Available"] = "Yes" if matched else "No"
        lead["Matched Company Name"]   = comp_norms.get(cn, "")

    # Count leads per company
    counts = {}
    for lead in leads:
        mn = lead["Matched Company Name"]
        if mn:
            counts[mn] = counts.get(mn, 0) + 1

    for comp in companies:
        comp["Lead Count"] = counts.get(comp["Company Name"], 0)

# ---------------------------------------------------------------------------
# WRITE LEADS SHEET
# ---------------------------------------------------------------------------

LEAD_COLS = [
    ("Source",               18),
    ("Full Name",            22),
    ("First Name",           16),
    ("Last Name",            16),
    ("Job Title",            28),
    ("Company",              28),
    ("Email",                30),
    ("Personal Email",       28),
    ("CC Email",             24),
    ("Email Status",         14),
    ("Phone",                18),
    ("LinkedIn",             38),
    ("Instagram",            30),
    ("Facebook",             30),
    ("Twitter",              26),
    ("Website",              28),
    ("Company Website",      28),
    ("Company Domain",       22),
    ("Industry",             26),
    ("City",                 16),
    ("State",                10),
    ("Country",              18),
    ("Current Stage",        18),
    ("Notion URL",           38),
    ("Instantly ID",         28),
    ("Company Data Available", 22),
    ("Matched Company Name", 28),
]

def write_leads_sheet(ws, leads):
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 32

    for col_idx, (col_name, col_w) in enumerate(LEAD_COLS, 1):
        c = ws.cell(row=1, column=col_idx, value=col_name)
        style_header(c)
        ws.column_dimensions[get_column_letter(col_idx)].width = col_w

    for row_idx, lead in enumerate(leads, 2):
        alt = (row_idx % 2 == 0)
        fills = {}
        # Colour the "Company Data Available" column
        cda_col = next(i for i, (n, _) in enumerate(LEAD_COLS, 1) if n == "Company Data Available")
        cda_val = lead.get("Company Data Available", "No")
        fills[cda_col] = YES_FILL if cda_val == "Yes" else NO_FILL

        # Email status colouring
        estat_col = next(i for i, (n, _) in enumerate(LEAD_COLS, 1) if n == "Email Status")
        estat_val = (lead.get("Email Status") or "").lower()
        if estat_val == "valid":
            fills[estat_col] = YES_FILL
        elif estat_val == "risky":
            fills[estat_col] = WARN_FILL

        row_vals = [lead.get(col_name, "") for col_name, _ in LEAD_COLS]
        write_row(ws, row_idx, row_vals, alt=alt, fills=fills)
        ws.row_dimensions[row_idx].height = 18

    ws.auto_filter.ref = f"A1:{get_column_letter(len(LEAD_COLS))}1"

# ---------------------------------------------------------------------------
# WRITE COMPANIES SHEET
# ---------------------------------------------------------------------------

COMP_COLS = [
    ("Company Name",        30),
    ("Website",             28),
    ("Core Services",       28),
    ("Positioning",         28),
    ("Why Luxury",          34),
    ("Known For",           28),
    ("Instagram Handle",    20),
    ("Engagement Level",    16),
    ("Reddit URL",          36),
    ("Buyer Pain Points",   40),
    ("AI Opportunity",      28),
    ("AI Fit Logic",        38),
    ("Personalized Opener", 44),
    ("Retainer Potential",  20),
    ("Why They Can Pay",    34),
    ("Lead Count",          12),
]

def write_companies_sheet(ws, companies):
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 32

    for col_idx, (col_name, col_w) in enumerate(COMP_COLS, 1):
        c = ws.cell(row=1, column=col_idx, value=col_name)
        style_header(c)
        ws.column_dimensions[get_column_letter(col_idx)].width = col_w

    lc_col = next(i for i, (n, _) in enumerate(COMP_COLS, 1) if n == "Lead Count")

    for row_idx, comp in enumerate(companies, 2):
        alt = (row_idx % 2 == 0)
        fills = {}
        lc = comp.get("Lead Count", 0)
        fills[lc_col] = YES_FILL if lc > 0 else NO_FILL

        row_vals = [comp.get(col_name, "") for col_name, _ in COMP_COLS]
        write_row(ws, row_idx, row_vals, alt=alt, fills=fills)
        ws.row_dimensions[row_idx].height = 20

    ws.auto_filter.ref = f"A1:{get_column_letter(len(COMP_COLS))}1"

# ---------------------------------------------------------------------------
# WRITE SUMMARY SHEET
# ---------------------------------------------------------------------------

def write_summary_sheet(ws, leads, companies):
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 18

    summary_data = [
        ("MASTER LEADS – SUMMARY", ""),
        ("", ""),
        ("Total Leads",                    len(leads)),
        ("Leads with Company Data",        sum(1 for l in leads if l.get("Company Data Available") == "Yes")),
        ("Leads without Company Data",     sum(1 for l in leads if l.get("Company Data Available") == "No")),
        ("", ""),
        ("Total Companies",                len(companies)),
        ("Companies with Lead(s)",         sum(1 for c in companies if c.get("Lead Count", 0) > 0)),
        ("Companies without Leads",        sum(1 for c in companies if c.get("Lead Count", 0) == 0)),
        ("", ""),
        ("Leads – Source: Lead Data",      sum(1 for l in leads if l.get("Source") == "Lead Data")),
        ("Leads – Source: Lead Set 1",     sum(1 for l in leads if l.get("Source") == "Lead Set 1")),
        ("", ""),
        ("Valid Email Leads",              sum(1 for l in leads if (l.get("Email Status") or "").lower() == "valid")),
        ("Risky Email Leads",              sum(1 for l in leads if (l.get("Email Status") or "").lower() == "risky")),
        ("No Email Status (Dubai leads)",  sum(1 for l in leads if not l.get("Email Status"))),
        ("", ""),
        ("Countries covered",              len({l.get("Country","") for l in leads if l.get("Country")})),
        ("UAE leads",                      sum(1 for l in leads if "arab" in (l.get("Country") or "").lower() or "dubai" in (l.get("City") or "").lower())),
        ("US leads",                       sum(1 for l in leads if "united states" in (l.get("Country") or "").lower())),
    ]

    for row_idx, (label, val) in enumerate(summary_data, 1):
        c_label = ws.cell(row=row_idx, column=1, value=label)
        c_val   = ws.cell(row=row_idx, column=2, value=val)
        if row_idx == 1:
            c_label.font = Font(name="Calibri", bold=True, size=14, color="1F4E79")
            c_val.font   = BOLD_FONT
        elif label == "":
            pass
        elif val == "":
            c_label.font = Font(name="Calibri", bold=True, size=11, color="2E75B6")
        else:
            c_label.font = BODY_FONT
            c_val.font   = BOLD_FONT
            c_val.alignment = CENTER

# ---------------------------------------------------------------------------
# MAIN
# ---------------------------------------------------------------------------

def main():
    out_path       = "C:/Users/Komal/Documents/AI Saas/AI Medical/data/master_leads.xlsx"
    lead_set1_path = "C:/Users/Komal/Documents/AI Saas/AI Medical/data/lead_set1.csv"

    print("Parsing CSV data...")
    lead_data_rows  = parse_csv_text(LEAD_DATA_CSV)
    company_rows    = parse_csv_text(COMPANY_DETAILS_CSV)

    # Read Lead Set 1 from disk (full dataset)
    import csv, os
    with open(lead_set1_path, encoding="utf-8-sig") as f:
        lead_set1_rows = list(csv.DictReader(f))
    # Filter out blank rows and any duplicate header rows embedded in the CSV
    lead_set1_rows = [
        r for r in lead_set1_rows
        if r.get("First Name", "").strip() not in ("", "First Name")
        and r.get("Email", "").strip() not in ("", "Email")
    ]

    print(f"  Lead Data rows  : {len(lead_data_rows)}")
    print(f"  Company rows    : {len(company_rows)}")
    print(f"  Lead Set1 rows  : {len(lead_set1_rows)}")

    leads     = build_leads(lead_data_rows, lead_set1_rows)
    companies = build_companies(company_rows)

    cross_reference(leads, companies)

    print(f"\nBuilding Excel...")
    print(f"  Leads           : {len(leads)}")
    print(f"  Companies       : {len(companies)}")
    print(f"  With company data: {sum(1 for l in leads if l.get('Company Data Available')=='Yes')}")

    wb = openpyxl.Workbook()

    # Summary sheet first
    ws_summary = wb.active
    ws_summary.title = "Summary"
    write_summary_sheet(ws_summary, leads, companies)

    # Leads sheet
    ws_leads = wb.create_sheet("Leads")
    write_leads_sheet(ws_leads, leads)

    # Companies sheet
    ws_companies = wb.create_sheet("Companies")
    write_companies_sheet(ws_companies, companies)

    wb.save(out_path)
    print(f"\nSaved to: {out_path}")

    # Quick verify
    wb2 = openpyxl.load_workbook(out_path)
    for sn in wb2.sheetnames:
        ws = wb2[sn]
        print(f"  Sheet '{sn}': {ws.max_row} rows x {ws.max_column} cols")

    # Delete intermediate CSV – only master_leads.xlsx should remain
    if os.path.exists(lead_set1_path):
        os.remove(lead_set1_path)
        print(f"\nDeleted: {lead_set1_path}")

if __name__ == "__main__":
    main()
